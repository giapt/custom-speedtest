import speedtest
from openpyxl import Workbook

servers = [11082,8156,6085,9903,6342,2552,5774,9174,11430,2428,11194,10308,7215,9594,16873,
19294,19060,6106,9994,12473,9331,12635,8158,18250,8491,15324,2515,3381,16749,9668,11342]
def getInfoFromServer(server):
    array = [server]
    obj = {}
    s = speedtest.Speedtest()
    s.get_servers(array)
    s.get_best_server()
    s.download()
    s.upload()
    s.results.share()

    results_dict = s.results.dict()

    client_info['lon'] = results_dict['client']['lon']
    client_info['lat'] =  results_dict['client']['lat']
    client_info['ip'] = results_dict['client']['ip']
    client_info['isp'] = results_dict['client']['isp']
    client_info['country'] = results_dict['client']['country']

    obj['ping'] = results_dict['ping']
    # print ping
    obj['download'] = round(results_dict['download'] / 8000000, 3)
    # print download
    obj['upload'] = round(results_dict['upload'] / 8000000, 3)
    obj['url'] = results_dict['share']
    # print upload

    #server
    obj['server_country'] = results_dict['server']['cc']
    obj['server_sponsor'] = results_dict['server']['sponsor']
    obj['server_id'] = results_dict['server']['id']
    obj['server_name'] = results_dict['server']['name']
    print obj['server_name'] , obj['server_sponsor']
    datas.append(obj)
    pass

client_info = {}
datas = []
for server in servers:
    getInfoFromServer(server)
    pass

wb = Workbook()
ws = wb.active
ws['A1'] = 'Client Info'
ws['A2'] = 'IP'
ws['B2'] = 'ISP'
ws['C2'] = 'Country'
ws['D2'] = 'Lon'
ws['E2'] = 'Lat'
ws['A3'] = client_info['ip']
ws['B3'] = client_info['isp']
ws['C3'] = client_info['country']
ws['D3'] = client_info['lon']
ws['E3'] = client_info['lat']

ws['A5'] = 'Server Id'
ws['B5'] = 'Server Name'
ws['C5'] = 'Ping'
ws['D5'] = 'Download (mb/s)'
ws['E5'] = 'Upload (mb/s)'
ws['F5'] = 'Server Country'
ws['G5'] = 'Url'
count = 0
for data in datas:
    count += 1
    ws['A' + str(5 + count)] = data['server_id']
    ws['B' + str(5 + count)] = data['server_name'] + ' - ' + data['server_sponsor']
    
    ws['C' + str(5 + count)] = data['ping']
    ws['D' + str(5 + count)] = str(data['download'])
    ws['E' + str(5 + count)] = str(data['upload'])
    ws['F' + str(5 + count)] = data['server_country']
    ws['G' + str(5 + count)] = data['url']
    pass

for col in ws.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = max_length + 1
    ws.column_dimensions[column].width = adjusted_width
    
wb.save("speedtest_server_vn.xlsx")